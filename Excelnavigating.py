import openpyxl
#to open a xl
workbook = openpyxl.load_workbook(r"C:\Users\SoulHunter V2.0\Desktop\Work\Workspace\Excels\Testdata.xlsx")

# to navigate to current sheet
sheet = workbook.active

# to change a value
sheet.cell(2,2).value = 'changed-again'
#workbook.save(r"C:\Users\SoulHunter V2.0\Desktop\Work\Workspace\Excels\Testdata.xlsx")
#print(sheet.cell(2,2).value)

print(sheet.max_row)
print(sheet.max_column)

# to get password for a username
for m in range(1, sheet.max_row+1):
    for n in range(1, sheet.max_column+1):
        if sheet.cell(m,n).value == '6yummy':
            print(sheet.cell(m,n-1).value)
